from aiogram import Bot,types
from aiogram.dispatcher import Dispatcher
from aiogram.utils import executor
from config import token
import keyboard2 as kb
import openpyxl

bot = Bot(token)
dp = Dispatcher(bot)
FIO = False
Phone = False
EMail = False
Work = False
Date = False
DATA = []
wb = openpyxl.load_workbook('work.xlsx')
sheet = wb['works']

async def print_all_works(ms):
    text = ''
    for i in range(2,sheet.max_row+1):
        text += sheet['A1'].value + ' : ' + sheet['A'+str(i)].value + '\n'
        text += sheet['B1'].value + ' : ' + sheet['B'+str(i)].value + '\n'
        text += sheet['C1'].value + ' : ' + sheet['C'+str(i)].value + '\n'
        text += sheet['D1'].value + ' : ' + sheet['D'+str(i)].value + '\n'
        text += sheet['E1'].value + ' : ' + sheet['E'+str(i)].value + '\n\n'
    await bot.send_message(ms.from_user.id,text)
    


def write_down_data(DATA):
    row = sheet.max_row + 1
    column = 'A'
    for cell in DATA:
        cor = column + str(row)
        sheet[cor] = cell
        column = chr(ord(column) + 1)
    wb.save('work.xlsx')  


@dp.message_handler(commands=['start'])
async def answer_start(ms:types.Message):
    if ms['from']['id'] == 755374812:
        text = 'Вы зашли как админ'
        await bot.send_message(ms.from_user.id,text,reply_markup = kb.keyboard)
    else:
        text = 'Вы зашли как исполнитель'
        await bot.send_message(ms.from_user.id,text,reply_markup = kb.keyboard3)
    text += str(ms['from'])
        
    

@dp.message_handler()
async def dialog_with_user(ms:types.Message):
    if ms['from']['id'] == 755374812:
        global FIO,Phone,EMail,Work,Date,DATA
        if ms.text == 'Оформить заказ':
            await bot.send_message(ms.from_user.id,'Введите ФИО заказчика:\nИванов Иван Иванович',reply_markup = kb.keyboard2)
            FIO = True
            return
        if ms.text == 'Отменить':
            await bot.send_message(ms.from_user.id,'ввод данных отменён',reply_markup = kb.keyboard)
            FIO = False
            Phone = False
            EMail = False
            Work = False
            Date = False
            DATA = []
            return
        if FIO:
            await bot.send_message(ms.from_user.id,'Введите номер телефона заказчика:\n79093601272')
            DATA.append(ms.text)
            FIO = False
            Phone = True
            return
        if Phone:
            await bot.send_message(ms.from_user.id,'Введите электронную почту заказчика:\nabc@mail\.ru')
            DATA.append(ms.text)
            Phone = False
            EMail = True
            return
        if EMail:
            await bot.send_message(ms.from_user.id,'Введите заказ:\nСварить ужин')
            DATA.append(ms.text)
            EMail = False
            Work = True
            return
        if Work:
            await bot.send_message(ms.from_user.id,'Введите дату заказа:\n17\.10\.2001')
            DATA.append(ms.text)
            Work = False
            Date = True
            return
        if Date:
            DATA.append(ms.text)
            write_down_data(DATA) 
            await bot.send_message(ms.from_user.id,'Данные сохранены',reply_markup = kb.keyboard)
            Date = False 
            return
        else:
            await bot.send_message(ms.from_user.id,'Выберите команду')    
    else:
        if ms.text == 'Показать заказы':
            await print_all_works(ms)
            await bot.send_document(ms.from_user.id,open('work.xlsx','rb'),caption = 'Список заказов')

    
if __name__ == '__main__':
    executor.start_polling(dp)
